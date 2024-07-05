#  任务分发
from openpyxl import load_workbook,Workbook

class Dispach:
    missions = []
    def __init__(self,file_path):
        # df = pd.read_excel(file_path)
  
        # 初始化表格
        workbook = load_workbook(file_path)
        self.workbook = workbook
        print(workbook.sheetnames,'表名')
        sheet = workbook.active
        print(sheet)
        for row in sheet.iter_rows(values_only=True):
            self.missions.append(row)
        print('加载完成')
    
    # def insert(self):


# path = './test.xlsx'
# Dispach(path)